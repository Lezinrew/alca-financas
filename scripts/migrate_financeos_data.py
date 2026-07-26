#!/usr/bin/env python3
"""
Migração Fase 1 (ADR-0008 do FinanceOS): TabelaTransacoes (Excel) -> alca-financas (Supabase).

Escopo: núcleo financeiro apenas (transações + categorias). Contas e regras
de categorização ficam fora desta Fase 1 (ver
specs/proposals/2026-07-26-migracao-financeos-supabase-fase1.md).

Uso:
    python scripts/migrate_financeos_data.py \
        --xlsm "caminho/para/controle_financeiro_maio2025_dashboard.xlsm" \
        --tenant-id <uuid> --user-id <uuid> [--dry-run]

--dry-run: faz toda a leitura, mapeamento e cálculo de dedup_key, mas não
escreve nada no Supabase. Sempre rodar --dry-run primeiro.

Pré-requisitos:
- Migration 20260726000004_transactions_legacy_id_source_file.sql aplicada
  (obrigatória para a carga real; --dry-run não precisa dela).
- .env na raiz do alca-financas com SUPABASE_URL e SUPABASE_SERVICE_ROLE_KEY.
- Exatamente 1 conta já cadastrada no tenant alvo (checagem de segurança —
  esta Fase 1 mapeia todas as transações para essa única conta, dado que a
  coluna Conta da planilha real está preenchida como "A definir").
"""
import argparse
import os
import sys
import uuid
from datetime import datetime, date as date_cls

import openpyxl
from openpyxl.utils.cell import range_boundaries
from dotenv import load_dotenv
from supabase import create_client

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))
from services.import_service import compute_dedup_key  # mesma fórmula do import da app

STATUS_MAP = {
    'pago': 'paid',
    'recebido': 'paid',
    'pendente': 'pending',
}
TYPE_MAP = {'despesa': 'expense', 'receita': 'income'}


def read_table(ws, table_name):
    tbl = ws.tables[table_name]
    min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
    headers = [ws.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)]
    rows = []
    for r in range(min_row + 1, max_row + 1):
        row = {h: ws.cell(row=r, column=min_col + i).value for i, h in enumerate(headers)}
        rows.append(row)
    return rows


def parse_rows(raw_rows):
    parsed, skipped = [], []
    for i, row in enumerate(raw_rows):
        legacy_id = row.get('ID')
        tipo = row.get('Tipo')
        data_val = row.get('Data')
        valor = row.get('Valor')
        if legacy_id is None or tipo is None or data_val is None or valor is None:
            skipped.append((i + 2, 'campos obrigatórios ausentes (ID/Tipo/Data/Valor)'))
            continue

        tx_type = TYPE_MAP.get(str(tipo).strip().lower())
        if not tx_type:
            skipped.append((i + 2, f'tipo desconhecido: {tipo!r}'))
            continue

        status_raw = str(row.get('Status') or '').strip().lower()
        status = STATUS_MAP.get(status_raw, 'pending')

        if hasattr(data_val, 'strftime'):
            date_str = data_val.strftime('%Y-%m-%d')
        elif isinstance(data_val, (date_cls, datetime)):
            date_str = data_val.isoformat()[:10]
        else:
            date_str = str(data_val)[:10]

        description = str(row.get('Descrição') or '').strip() or 'Sem descrição'
        resp_raw = row.get('Responsável')
        responsible_person = str(resp_raw).strip() if resp_raw else None

        parsed.append({
            'legacy_id': int(legacy_id),
            'date': date_str,
            'type': tx_type,
            'amount': abs(float(valor)),
            'description': description,
            'category_name': str(row.get('Categoria') or '').strip() or 'Outros',
            'responsible_person': responsible_person,
            'status': status,
            'source_file': (str(row.get('Arquivo Origem') or '').strip() or None),
        })
    return parsed, skipped


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('--xlsm', required=True, help='Caminho para o .xlsm de produção do FinanceOS')
    parser.add_argument('--tenant-id', required=True)
    parser.add_argument('--user-id', required=True)
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    load_dotenv()
    supabase_url = os.environ.get('SUPABASE_URL')
    service_key = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')
    if not supabase_url or not service_key:
        print('ERRO: SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY não definidos (.env).')
        sys.exit(1)
    sb = create_client(supabase_url, service_key)

    print(f'=== Lendo {args.xlsm} ===')
    wb = openpyxl.load_workbook(args.xlsm, data_only=True, keep_vba=False)
    ws = wb['Transações']
    raw_rows = read_table(ws, 'TabelaTransacoes')
    parsed, skipped = parse_rows(raw_rows)

    print(f'Linhas na planilha: {len(raw_rows)} | válidas: {len(parsed)} | ignoradas: {len(skipped)}')
    for line_num, reason in skipped:
        print(f'  - linha {line_num}: {reason}')

    print(f'\n=== Resolvendo conta de destino (tenant {args.tenant_id}) ===')
    accounts = sb.table('accounts').select('id,name').eq('tenant_id', args.tenant_id).execute().data
    if len(accounts) != 1:
        print(f'ERRO: esperava exatamente 1 conta no tenant, encontrei {len(accounts)}. Abortando '
              f'(esta Fase 1 assume conta única — ver docstring).')
        sys.exit(1)
    account_id = accounts[0]['id']
    print(f'Conta de destino: {accounts[0]["name"]} ({account_id})')

    print(f'\n=== Resolvendo categorias ===')
    existing_cats = sb.table('categories').select('id,name,type').eq('tenant_id', args.tenant_id).execute().data
    cat_map = {(c['name'].strip().lower(), c['type']): c['id'] for c in existing_cats}

    seen_new = {}
    for tx in parsed:
        key = (tx['category_name'].strip().lower(), tx['type'])
        if key not in cat_map and key not in seen_new:
            seen_new[key] = {'name': tx['category_name'], 'type': tx['type']}

    matched_count = sum(
        1 for tx in parsed if (tx['category_name'].strip().lower(), tx['type']) in cat_map
    )
    print(f'Categorias já existentes reutilizadas: {matched_count} transações')
    print(f'Categorias novas a criar: {len(seen_new)}')
    for (name_lower, ctype), info in seen_new.items():
        print(f'  - {info["name"]!r} ({ctype})')

    if not args.dry_run:
        for key, info in seen_new.items():
            created = sb.table('categories').insert({
                'user_id': args.user_id,
                'tenant_id': args.tenant_id,
                'name': info['name'],
                'type': info['type'],
                'color': '#6C757D',
                'icon': 'circle',
            }).execute().data[0]
            cat_map[key] = created['id']

    print(f'\n=== Deduplicação (contra {len(parsed)} novas + histórico existente) ===')
    existing_txs = (
        sb.table('transactions')
        .select('date,amount,description,account_id,type')
        .eq('tenant_id', args.tenant_id)
        .execute()
        .data
    )
    existing_keys = {
        compute_dedup_key(t['date'], t['amount'], t['description'], t['account_id'], t['type'])
        for t in existing_txs
    }
    print(f'Transações já existentes no tenant: {len(existing_txs)}')

    to_insert = []
    duplicates = []
    seen_in_batch = set()
    for tx in parsed:
        key = compute_dedup_key(tx['date'], tx['amount'], tx['description'], account_id, tx['type'])
        if key in existing_keys or key in seen_in_batch:
            duplicates.append(tx)
            continue
        seen_in_batch.add(key)
        cat_id = cat_map.get((tx['category_name'].strip().lower(), tx['type']))
        to_insert.append({
            'id': str(uuid.uuid4()),
            'user_id': args.user_id,
            'tenant_id': args.tenant_id,
            'account_tenant_id': args.tenant_id,
            'category_tenant_id': args.tenant_id,
            'description': tx['description'],
            'amount': tx['amount'],
            'type': tx['type'],
            'category_id': cat_id,
            'account_id': account_id,
            'date': tx['date'],
            'is_recurring': False,
            'status': tx['status'],
            'responsible_person': tx['responsible_person'],
            'entry_source': 'manual',
            'legacy_id': tx['legacy_id'],
            'source_file': tx['source_file'],
            'dedup_key': key,
        })

    print(f'Duplicatas (já existem no Supabase, puladas): {len(duplicates)}')
    for d in duplicates:
        print(f'  - legacy_id={d["legacy_id"]} data={d["date"]} valor={d["amount"]}')
    print(f'Transações novas a inserir: {len(to_insert)}')

    if args.dry_run:
        print('\nDRY-RUN: nenhuma escrita realizada no Supabase.')
        return

    if to_insert:
        sb.table('transactions').insert(to_insert).execute()
        print(f'\n{len(to_insert)} transações inseridas com sucesso no tenant {args.tenant_id}.')
    else:
        print('\nNada a inserir (tudo já existia ou foi ignorado).')


if __name__ == '__main__':
    main()
